"""
TaxEasy Global - 해외 양도소득세 PDF 파싱 & 홈택스 엑셀 변환 엔진
=================================================================
증권사 PDF → 거래내역 파싱 → 세금 계산 → 홈택스 업로드용 엑셀 생성

지원 증권사: 한국투자증권 (향후 키움, 미래에셋, 삼성 추가 예정)
지원 파일: PDF, Excel(XLSX/XLS), CSV
"""

import re
import sys
import json
from datetime import datetime, timedelta
from pathlib import Path

import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ============================================================
# 1. PDF PARSER - 증권사별 거래내역 PDF 파싱
# ============================================================

class KoreaInvestmentPDFParser:
    """한국투자증권 해외주식 양도소득금액 계산내역 PDF 파서"""

    # 거래내역 테이블 헤더 패턴 (11개 컬럼)
    TRADE_HEADERS = ['주식', '주식', '주식', '양도', '양도', '주당', '양도가액', '주당', '취득가액', '제비용', '비용차감']

    def parse(self, pdf_path: str) -> dict:
        """PDF 파일을 파싱하여 거래내역과 메타정보를 반환"""
        result = {
            'broker': '한국투자증권',
            'meta': {},
            'trades': [],
            'summary_by_stock': [],
            'totals': {},
        }

        with pdfplumber.open(pdf_path) as pdf:
            all_trade_rows = []
            all_summary_rows = []

            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    if not table or len(table) < 2:
                        continue

                    first_row = table[0]
                    cols = len(first_row) if first_row else 0

                    # 메타 정보 테이블 (4열, 성명/주민번호 등)
                    if cols == 4 and any('성명' in str(c) for c in first_row if c):
                        result['meta'] = self._parse_meta(table)

                    # 거래 상세 테이블 (11열)
                    elif cols == 11:
                        is_header = any('종목명' in str(c) for c in first_row if c)
                        rows = table[1:] if is_header else table
                        for row in rows:
                            parsed = self._parse_trade_row(row)
                            if parsed:
                                all_trade_rows.append(parsed)

                    # 종목별 요약 테이블 (10열)
                    elif cols == 10:
                        is_header = any('종목명' in str(c) for c in first_row if c)
                        rows = table[1:] if is_header else table
                        for row in rows:
                            parsed = self._parse_summary_row(row)
                            if parsed:
                                all_summary_rows.append(parsed)

                    # 비용 테이블 (2열)
                    elif cols == 2 and any('수수료' in str(c) for c in first_row if c):
                        result['extra_costs'] = self._parse_costs(table)

            result['trades'] = all_trade_rows
            result['summary_by_stock'] = all_summary_rows

            # 합계 계산
            if all_trade_rows:
                result['totals'] = {
                    'total_sell': sum(t['sell_total'] for t in all_trade_rows),
                    'total_buy': sum(t['buy_total'] for t in all_trade_rows),
                    'total_expenses': sum(t['expenses'] for t in all_trade_rows),
                    'total_profit_loss': sum(t['profit_loss'] for t in all_trade_rows),
                    'trade_count': len(all_trade_rows),
                }

        return result

    def _clean_val(self, v):
        """셀 값에서 줄바꿈 제거 및 정리"""
        if v is None:
            return ''
        return str(v).replace('\n', '').strip()

    def _parse_num(self, v):
        """숫자 문자열 파싱 (쉼표, 줄바꿈 제거)"""
        s = self._clean_val(v)
        if not s or s == '-':
            return 0
        s = s.replace(',', '')
        try:
            return float(s)
        except ValueError:
            return 0

    def _parse_meta(self, table):
        """메타 정보 테이블 파싱"""
        meta = {}
        for row in table:
            for i, cell in enumerate(row):
                s = str(cell) if cell else ''
                if '성명' in s:
                    meta['name'] = str(row[i + 1]) if i + 1 < len(row) and row[i + 1] else ''
                if '금융기관명' in s:
                    meta['broker'] = str(row[i + 1]) if i + 1 < len(row) and row[i + 1] else ''
                if '계좌번호' in s:
                    meta['account'] = str(row[i + 1]) if i + 1 < len(row) and row[i + 1] else ''
        return meta

    def _parse_trade_row(self, row):
        """개별 거래 행 파싱 (11열 테이블)"""
        if len(row) < 11:
            return None

        stock_name = self._clean_val(row[0])
        if not stock_name or '합' in stock_name:
            return None

        stock_code = self._clean_val(row[1])
        stock_type = self._clean_val(row[2])
        shares = self._parse_num(row[3])
        sell_date = self._clean_val(row[4]).replace('.', '-')
        sell_pps = self._parse_num(row[5])
        sell_total = self._parse_num(row[6])
        buy_pps = self._parse_num(row[7])
        buy_total = self._parse_num(row[8])
        expenses = self._parse_num(row[9])
        profit_loss = self._parse_num(row[10])

        if shares <= 0 and sell_total <= 0:
            return None

        return {
            'stock_name': stock_name,
            'stock_code': stock_code,
            'stock_type_code': stock_type or '61',
            'shares': shares,
            'sell_date': sell_date,
            'sell_price_per_share': sell_pps,
            'sell_total': int(round(sell_total)),
            'buy_price_per_share': buy_pps,
            'buy_total': int(round(buy_total)),
            'expenses': int(round(expenses)),
            'profit_loss': int(round(profit_loss)),
            'country_code': self._detect_country(stock_code),
        }

    def _parse_summary_row(self, row):
        """종목별 요약 행 파싱"""
        if len(row) < 10:
            return None
        name = self._clean_val(row[0])
        if not name or name == '합 계':
            return None
        return {
            'stock_name': name,
            'stock_code': self._clean_val(row[1]),
            'total_shares': self._parse_num(row[3]),
            'avg_sell_pps': self._parse_num(row[4]),
            'total_sell': self._parse_num(row[5]),
            'avg_buy_pps': self._parse_num(row[6]),
            'total_buy': self._parse_num(row[7]),
            'total_expenses': self._parse_num(row[8]),
            'total_profit_loss': self._parse_num(row[9]),
        }

    def _parse_costs(self, table):
        """추가 비용 테이블 파싱"""
        costs = {}
        for row in table:
            label = self._clean_val(row[0])
            value = self._parse_num(row[1]) if len(row) > 1 else 0
            if label:
                costs[label] = value
        return costs

    def _detect_country(self, code):
        """ISIN 코드에서 국가코드 추출"""
        code = code.upper()
        if code.startswith('US'):
            return 'US'
        elif code.startswith('KY'):
            return 'US'  # 케이만제도 등록이지만 미국시장 거래
        elif code.startswith('JP'):
            return 'JP'
        elif code.startswith('HK'):
            return 'HK'
        elif code.startswith('GB'):
            return 'GB'
        elif code.startswith('DE'):
            return 'DE'
        return 'US'


# ============================================================
# 2. MULTI-FORMAT PARSER - Excel/CSV 파싱
# ============================================================

class ExcelCSVParser:
    """엑셀 및 CSV 거래내역 파서"""

    # 가능한 컬럼명 매핑
    COL_MAPS = {
        'stock_name': ['주식 종목명', '주식종목명', '종목명', '종목', 'stock_name'],
        'stock_code': ['주식종목코드', '종목코드', 'ISIN', '국제증권식별번호', 'stock_code'],
        'shares': ['양도주식수', '양도주식 수', '취득유형별 양도주식 수', '수량', 'shares'],
        'sell_date': ['양도일자', '매도일자', '매도일', 'sell_date'],
        'sell_pps': ['주당양도가액', '주당매도가액', '매도단가', 'sell_price_per_share'],
        'sell_total': ['양도가액', '매도금액', 'sell_total'],
        'buy_date': ['취득일자', '매수일자', '매수일', 'buy_date'],
        'buy_pps': ['주당취득가액', '주당매수가액', '매수단가', 'buy_price_per_share'],
        'buy_total': ['취득가액', '매수금액', 'buy_total'],
        'expenses': ['제비용', '필요경비', '수수료', 'expenses'],
        'country': ['국외자산국가코드', '국가코드', 'country'],
    }

    def parse(self, file_path: str) -> list:
        """Excel 또는 CSV 파일을 파싱하여 거래 리스트 반환"""
        ext = Path(file_path).suffix.lower()
        if ext in ('.xlsx', '.xls'):
            df = pd.read_excel(file_path, header=0)
        elif ext in ('.csv', '.tsv'):
            sep = '\t' if ext == '.tsv' else ','
            df = pd.read_csv(file_path, sep=sep, header=0)
        else:
            raise ValueError(f"지원하지 않는 파일 형식: {ext}")

        col_map = self._detect_columns(df.columns)
        trades = []
        for _, row in df.iterrows():
            trade = self._map_row(row, col_map)
            if trade and trade['shares'] > 0:
                trades.append(trade)
        return trades

    def _detect_columns(self, columns):
        """데이터프레임 컬럼을 내부 필드명으로 매핑"""
        mapping = {}
        cols_clean = {str(c).replace('\n', ' ').strip(): c for c in columns}

        for field, candidates in self.COL_MAPS.items():
            for cand in candidates:
                for clean_name, orig_name in cols_clean.items():
                    if cand in clean_name or clean_name in cand:
                        mapping[field] = orig_name
                        break
                if field in mapping:
                    break
        return mapping

    def _map_row(self, row, col_map):
        """한 행을 거래 딕셔너리로 변환"""
        def get(field, default=''):
            if field in col_map:
                v = row.get(col_map[field], default)
                return v if pd.notna(v) else default
            return default

        def get_num(field):
            v = get(field, 0)
            if isinstance(v, (int, float)):
                return float(v)
            return float(str(v).replace(',', '') or 0)

        def get_date(field):
            v = get(field, '')
            if isinstance(v, datetime):
                return v.strftime('%Y-%m-%d')
            s = str(v).replace('/', '-').strip()
            if len(s) >= 10:
                return s[:10]
            return s

        name = str(get('stock_name'))
        if not name:
            return None

        sell_total = int(round(get_num('sell_total')))
        buy_total = int(round(get_num('buy_total')))
        expenses = int(round(get_num('expenses')))

        return {
            'stock_name': name,
            'stock_code': str(get('stock_code')),
            'stock_type_code': '61',
            'shares': get_num('shares'),
            'sell_date': get_date('sell_date'),
            'sell_price_per_share': get_num('sell_pps'),
            'sell_total': sell_total,
            'buy_date': get_date('buy_date'),
            'buy_price_per_share': get_num('buy_pps'),
            'buy_total': buy_total,
            'expenses': expenses,
            'profit_loss': sell_total - buy_total - expenses,
            'country_code': str(get('country', 'US')) or 'US',
        }


# ============================================================
# 3. BUY DATE RESOLVER - 취득일자 추정/매핑
# ============================================================

class BuyDateResolver:
    """
    한국투자증권 PDF에는 취득일자가 없으므로,
    별도 매수내역 또는 사용자 입력으로 취득일자를 매핑.
    """

    def __init__(self):
        self.buy_dates = {}  # {stock_code: buy_date}

    def load_from_reference(self, ref_file: str):
        """최종 결과 파일이나 매수내역 파일에서 취득일자 추출"""
        try:
            df = pd.read_excel(ref_file, header=0)
            for _, row in df.iterrows():
                code_col = None
                date_col = None
                for c in df.columns:
                    cn = str(c).replace('\n', '')
                    if 'ISIN' in cn or '종목코드' in cn or '국제증권식별번호' in cn:
                        code_col = c
                    if '취득일자' in cn:
                        date_col = c
                if code_col and date_col:
                    code = str(row[code_col]).strip()
                    date = row[date_col]
                    if pd.notna(date) and code:
                        if isinstance(date, datetime):
                            self.buy_dates[code] = date.strftime('%Y-%m-%d')
                        else:
                            self.buy_dates[code] = str(date)[:10]
        except Exception as e:
            print(f"참조 파일 로드 실패: {e}")

    def load_manual(self, mapping: dict):
        """수동 매핑 {stock_code: buy_date}"""
        self.buy_dates.update(mapping)

    def resolve(self, trade: dict) -> str:
        """거래에 대한 취득일자를 반환"""
        if trade.get('buy_date'):
            return trade['buy_date']
        code = trade.get('stock_code', '')
        return self.buy_dates.get(code, '')


# ============================================================
# 4. TAX CALCULATOR - 양도소득세 계산 엔진
# ============================================================

class TaxCalculator:
    """해외주식 양도소득세 계산기"""

    BASIC_DEDUCTION = 2_500_000  # 기본공제 250만원
    TAX_RATE = 0.22             # 22% (국세 20% + 지방세 2%)

    def calculate(self, trades: list) -> dict:
        """거래 리스트에서 세금 계산 결과를 반환"""
        total_sell = sum(t['sell_total'] for t in trades)
        total_buy = sum(t['buy_total'] for t in trades)
        total_expenses = sum(t['expenses'] for t in trades)
        total_pl = sum(t['profit_loss'] for t in trades)

        taxable = max(0, total_pl - self.BASIC_DEDUCTION)
        tax = int(round(taxable * self.TAX_RATE))

        profit_trades = [t for t in trades if t['profit_loss'] > 0]
        loss_trades = [t for t in trades if t['profit_loss'] < 0]

        return {
            'trade_count': len(trades),
            'total_sell': total_sell,
            'total_buy': total_buy,
            'total_expenses': total_expenses,
            'gross_profit_loss': total_pl,
            'basic_deduction': self.BASIC_DEDUCTION,
            'taxable_income': taxable,
            'tax_rate': self.TAX_RATE,
            'tax_amount': tax,
            'national_tax': int(round(taxable * 0.20)),
            'local_tax': int(round(taxable * 0.02)),
            'profit_trade_count': len(profit_trades),
            'loss_trade_count': len(loss_trades),
            'total_profit': sum(t['profit_loss'] for t in profit_trades),
            'total_loss': sum(t['profit_loss'] for t in loss_trades),
            'filing_deadline': '2025-05-31',
        }


# ============================================================
# 5. HOMETAX EXCEL GENERATOR - 홈택스 업로드용 엑셀 생성
# ============================================================

class HometaxExcelGenerator:
    """홈택스 주식 양도소득세 엑셀업로드 양식 생성기
    공식 템플릿(주식_엑셀업로드_양식.xlsx)을 베이스로 자료 시트에 데이터 주입
    """

    # 템플릿 경로 (taxeasy_engine.py 기준 상대 경로)
    TEMPLATE_PATH = Path(__file__).parent.parent.parent / '양도소득세' / '주식_엑셀업로드_양식.xlsx'

    def generate(self, trades: list, output_path: str, buy_date_resolver=None):
        """공식 템플릿에 거래 데이터를 주입하여 홈택스 업로드용 엑셀 생성"""
        from openpyxl import load_workbook

        # 공식 템플릿 로드 (4개 시트 + 스타일 + 인쇄 설정 모두 보존)
        template_path = self._find_template()
        wb = load_workbook(template_path)
        ws = wb['자료']

        # 기존 샘플 데이터 행 삭제 (1행 헤더 유지, 2행 이후 삭제)
        while ws.max_row > 1:
            ws.delete_rows(ws.max_row)

        # 데이터 행 작성
        for row_idx, trade in enumerate(trades, 2):
            buy_date = ''
            if buy_date_resolver:
                buy_date = buy_date_resolver.resolve(trade)
            elif trade.get('buy_date'):
                buy_date = trade['buy_date']

            row = [
                trade['stock_name'] or '',                       # A: 주식 종목명
                '',                                              # B: 사업자등록번호
                2,                                               # C: 국내/국외 구분 (숫자, 2=국외)
                trade['shares'],                                 # D: 취득유형별 양도주식 수
                '61',                                            # E: 세율구분 (문자열 코드)
                '61',                                            # F: 주식등 종류 (문자열 코드)
                '10',                                            # G: 양도물건 종류 (문자열 코드)
                '01',                                            # H: 취득유형 (선행0 유지)
                trade['sell_date'],                              # I: 양도일자 (YYYY-MM-DD)
                int(round(trade['sell_price_per_share'])),       # J: 주당양도가액
                int(round(trade['sell_total'])),                 # K: 양도가액
                buy_date,                                        # L: 취득일자 (YYYY-MM-DD)
                int(round(trade['buy_price_per_share'])),        # M: 주당취득가액
                int(round(trade['buy_total'])),                  # N: 취득가액
                int(round(trade['expenses'])),                   # O: 필요경비
                '', '', '', '', '',                              # P-T: 빈값
                trade.get('stock_code', '') or '',               # U: ISIN/종목코드
                trade.get('country_code', 'US'),                 # V: 국가코드
                '',                                              # W: 국외자산내용
            ]
            for col_idx, val in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        wb.save(output_path)
        return output_path

    def _find_template(self) -> Path:
        """템플릿 파일 경로 찾기 (NFD 파일시스템 대응)"""
        import unicodedata
        # 직접 경로 시도
        if self.TEMPLATE_PATH.exists():
            return self.TEMPLATE_PATH

        # NFD 인코딩된 파일시스템 대응: bytes로 탐색
        target_name = '주식_엑셀업로드_양식.xlsx'
        parent_dir = self.TEMPLATE_PATH.parent

        try:
            raw_parent = str(parent_dir).encode('utf-8')
            for entry in os.listdir(raw_parent):
                nfc = unicodedata.normalize('NFC', entry.decode('utf-8', errors='replace'))
                if target_name in nfc:
                    full = raw_parent + b'/' + entry
                    # 임시 복사본으로 반환
                    import tempfile, shutil
                    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
                    tmp.close()
                    with open(full, 'rb') as src, open(tmp.name, 'wb') as dst:
                        dst.write(src.read())
                    return Path(tmp.name)
        except Exception:
            pass

        raise FileNotFoundError(f'홈택스 템플릿 파일을 찾을 수 없습니다: {self.TEMPLATE_PATH}')


# ============================================================
# 6. MAIN PIPELINE - 전체 파이프라인
# ============================================================

class TaxEasyPipeline:
    """PDF → 파싱 → 세금계산 → 홈택스 엑셀 생성 전체 파이프라인"""

    def __init__(self):
        self.pdf_parser = KoreaInvestmentPDFParser()
        self.excel_parser = ExcelCSVParser()
        self.tax_calc = TaxCalculator()
        self.excel_gen = HometaxExcelGenerator()
        self.buy_date_resolver = BuyDateResolver()

    def process_pdf(self, pdf_path: str, buy_dates: dict = None) -> dict:
        """PDF 파일을 처리하여 파싱 결과 반환"""
        result = self.pdf_parser.parse(pdf_path)
        if buy_dates:
            self.buy_date_resolver.load_manual(buy_dates)
        return result

    def process_excel_csv(self, file_path: str) -> list:
        """Excel/CSV 파일을 파싱하여 거래 리스트 반환"""
        return self.excel_parser.parse(file_path)

    def calculate_tax(self, trades: list) -> dict:
        """세금 계산"""
        return self.tax_calc.calculate(trades)

    def generate_hometax_excel(self, trades: list, output_path: str) -> str:
        """홈택스 업로드용 엑셀 생성"""
        return self.excel_gen.generate(trades, output_path, self.buy_date_resolver)

    def run(self, input_path: str, output_path: str, buy_dates: dict = None):
        """전체 파이프라인 실행"""
        ext = Path(input_path).suffix.lower()

        if ext == '.pdf':
            result = self.process_pdf(input_path, buy_dates)
            trades = result['trades']
        else:
            trades = self.process_excel_csv(input_path)

        tax = self.calculate_tax(trades)
        self.generate_hometax_excel(trades, output_path)

        return {
            'trades': trades,
            'tax': tax,
            'output_file': output_path,
            'trade_count': len(trades),
        }


# ============================================================
# CLI 실행
# ============================================================

def print_report(result):
    """세금 계산 결과 출력"""
    tax = result['tax']
    print("\n" + "=" * 60)
    print("  TaxEasy Global - 해외 양도소득세 계산 결과")
    print("=" * 60)
    print(f"  거래 건수:        {tax['trade_count']}건")
    print(f"    - 이익 거래:    {tax['profit_trade_count']}건 (+{tax['total_profit']:,}원)")
    print(f"    - 손실 거래:    {tax['loss_trade_count']}건 ({tax['total_loss']:,}원)")
    print(f"  ────────────────────────────────────")
    print(f"  총 양도가액:      {tax['total_sell']:>15,}원")
    print(f"  총 취득가액:      {tax['total_buy']:>15,}원")
    print(f"  필요경비:         {tax['total_expenses']:>15,}원")
    print(f"  ────────────────────────────────────")
    print(f"  양도차익:         {tax['gross_profit_loss']:>15,}원")
    print(f"  기본공제:        -{tax['basic_deduction']:>14,}원")
    print(f"  ────────────────────────────────────")
    print(f"  과세표준:         {tax['taxable_income']:>15,}원")
    print(f"  세율:             {tax['tax_rate']*100:.0f}%")
    print(f"  ════════════════════════════════════")
    print(f"  ★ 납부할 세액:   {tax['tax_amount']:>15,}원")
    print(f"    (국세 {tax['national_tax']:,}원 + 지방세 {tax['local_tax']:,}원)")
    print(f"  ════════════════════════════════════")
    print(f"  신고 기한:        {tax['filing_deadline']}")
    print(f"  생성 파일:        {result['output_file']}")
    print()


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("사용법: python taxeasy_engine.py <PDF파일> [출력파일.xlsx]")
        print("예시:   python taxeasy_engine.py '양도소득세 2024data.pdf'")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else '홈택스_양도소득세_업로드용.xlsx'

    # 한국투자증권 PDF에는 취득일자가 없으므로 수동 매핑
    # (실제 서비스에서는 사용자 입력 또는 증권사 API로 해결)
    KNOWN_BUY_DATES = {
        'KYG4124C1096': '2023-06-22',   # 그랩 홀딩스
        'US7134481081': '2023-11-15',    # 펩시코
        'US83406F1021': '2023-12-26',    # 소파이 테크놀로지
    }

    pipeline = TaxEasyPipeline()
    pipeline.buy_date_resolver.load_manual(KNOWN_BUY_DATES)

    result = pipeline.run(input_file, output_file, KNOWN_BUY_DATES)
    print_report(result)
    print(f"✅ 홈택스 업로드용 엑셀 파일 생성 완료: {output_file}")
