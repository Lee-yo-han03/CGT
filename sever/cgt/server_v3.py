"""
TaxEasy Global - 통합 Backend API Server v3.0
==============================================
- 증권사 자동 감지 (선택 단계 제거)
- 다중 파일 업로드 지원
- 취득일자 자동채움 옵션
- 홈택스 신고 안내
"""

import json
import os
import re
import uuid
import traceback
import tempfile
import sys
import socketserver
from http.server import HTTPServer, SimpleHTTPRequestHandler
from urllib.parse import urlparse
from pathlib import Path
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from parsers import AutoDetectParser, GenericExcelParser

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# Tax Calculator
# ============================================================

BASIC_DEDUCTION = 2_500_000
TAX_RATE_NATIONAL = 0.20
TAX_RATE_LOCAL = 0.02
TAX_RATE_TOTAL = 0.22

def calculate_tax(trades):
    total_sell = sum(t.get('sell_total', 0) for t in trades)
    total_buy = sum(t.get('buy_total', 0) for t in trades)
    total_exp = sum(t.get('expenses', 0) for t in trades)
    total_pl = sum(t.get('profit_loss', 0) for t in trades)
    taxable = max(0, total_pl - BASIC_DEDUCTION)
    tax = int(round(taxable * TAX_RATE_TOTAL))
    return {
        'trade_count': len(trades),
        'total_sell': total_sell,
        'total_buy': total_buy,
        'total_expenses': total_exp,
        'gross_profit_loss': total_pl,
        'basic_deduction': BASIC_DEDUCTION,
        'taxable_income': taxable,
        'tax_rate': TAX_RATE_TOTAL,
        'tax_amount': tax,
        'national_tax': int(round(taxable * TAX_RATE_NATIONAL)),
        'local_tax': int(round(taxable * TAX_RATE_LOCAL)),
        'profit_trades': len([t for t in trades if t.get('profit_loss', 0) > 0]),
        'loss_trades': len([t for t in trades if t.get('profit_loss', 0) < 0]),
        'total_profit': sum(t.get('profit_loss', 0) for t in trades if t.get('profit_loss', 0) > 0),
        'total_loss': sum(t.get('profit_loss', 0) for t in trades if t.get('profit_loss', 0) < 0),
    }


# ============================================================
# Buy Date Auto-Fill Logic
# ============================================================

def auto_fill_buy_dates(trades):
    """
    취득일자 자동채움:
    연간 합산 신고 시 같은 연도 내 임의 날짜 가능.
    각 거래의 양도일 연도의 1월 1일을 취득일로 설정.
    조건: 취득일 < 양도일
    """
    filled = {}
    for t in trades:
        sell_date = t.get('sell_date', '')
        if sell_date and len(sell_date) >= 4:
            year = sell_date[:4]
            filled_date = f"{year}-01-01"
            # 취득일 < 양도일 확인
            if filled_date < sell_date:
                t['buy_date'] = filled_date
            else:
                # 같은 날이면 하루 전
                t['buy_date'] = f"{int(year)-1}-12-31"
    return trades


# ============================================================
# HomeTax Excel Generator
# ============================================================

HOMETAX_HEADERS = [
    '주식 종목명','사업자등록번호','국내/국외 구분','취득유형별\n양도주식 수',
    '세율구분','주식등 종류','양도물건 종류','취득유형','양도일자','주당양도가액',
    '양도가액','취득일자','주당취득가액','취득가액','필요경비','비과세\n양도소득금액',
    '감면종류','감면율','감면소득금액','과세이연여부',
    '국제증권식별번호\n(ISIN코드, 종목코드)','국외자산국가코드','국외자산내용',
]
COL_WIDTHS = [16,14,12,16,10,10,12,10,14,14,14,14,14,14,12,16,10,8,14,12,20,16,14]

def generate_hometax_excel(trades, buy_dates=None, auto_fill=False):
    """홈택스 엑셀 생성, 바이트 반환"""
    buy_dates = buy_dates or {}

    if auto_fill:
        trades = auto_fill_buy_dates([dict(t) for t in trades])

    wb = Workbook()
    ws = wb.active
    ws.title = '자료'

    hf = Font(bold=True, size=10, name='Arial')
    hfill = PatternFill('solid', fgColor='D9E2F3')
    ha = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(left=Side('thin'), right=Side('thin'),
                    top=Side('thin'), bottom=Side('thin'))
    df = Font(size=10, name='Arial')
    da = Alignment(vertical='center')
    na = Alignment(horizontal='right', vertical='center')

    for col, h in enumerate(HOMETAX_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font, cell.fill, cell.alignment, cell.border = hf, hfill, ha, border
    ws.row_dimensions[1].height = 40
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for ri, t in enumerate(trades, 2):
        code = t.get('stock_code', '')
        # buy_date 우선순위: buy_dates dict > trade 자체 > 비워두기
        bd = buy_dates.get(code, '') or t.get('buy_date', '')
        sell_pps = t.get('sell_price_per_share', 0)
        buy_pps = t.get('buy_price_per_share', 0)

        vals = [
            t.get('stock_name', ''), None, 2, t.get('shares', 0),
            61, 61, 10, '01',
            t.get('sell_date', ''),
            int(round(sell_pps)) if sell_pps else 0,
            t.get('sell_total', 0),
            bd,  # 취득일자 (자동채움 or 비워두기)
            int(round(buy_pps)) if buy_pps else 0,
            t.get('buy_total', 0),
            t.get('expenses', 0),
            None, None, None, None, None,
            code, t.get('country_code', 'US'), None,
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=col, value=val)
            cell.font, cell.border = df, border
            if isinstance(val, (int, float)) and val is not None:
                cell.alignment = na
                cell.number_format = '#,##0'
            else:
                cell.alignment = da

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(tmp.name)
    with open(tmp.name, 'rb') as f:
        data = f.read()
    os.unlink(tmp.name)
    return data


# ============================================================
# Session Storage (최대 1시간 보관, 최대 100개)
# ============================================================

sessions = {}
SESSION_TTL = 3600  # 1시간

def cleanup_sessions():
    """만료된 세션 정리"""
    now = datetime.now()
    expired = [sid for sid, s in sessions.items()
               if (now - datetime.fromisoformat(s['created'])).seconds > SESSION_TTL]
    for sid in expired:
        del sessions[sid]
    # 최대 100개 초과 시 가장 오래된 것 제거
    if len(sessions) > 100:
        oldest = sorted(sessions.items(), key=lambda x: x[1]['created'])
        for sid, _ in oldest[:len(sessions)-100]:
            del sessions[sid]


# ============================================================
# API Handler
# ============================================================

class TaxEasyHandler(SimpleHTTPRequestHandler):

    UPLOAD_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
    # 메인 index.html은 프로젝트 루트에 위치
    STATIC_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '..')

    def do_GET(self):
        parsed = urlparse(self.path)

        if parsed.path == '/api/health':
            self._json_response({
                'status': 'ok', 'version': '3.0.0',
                'sessions': len(sessions),
            })
            return

        if parsed.path == '/' or parsed.path == '/index.html':
            self._serve_file('index.html', 'text/html; charset=utf-8')
        else:
            super().do_GET()

    def do_POST(self):
        parsed = urlparse(self.path)
        try:
            if parsed.path == '/api/upload':
                self._handle_upload()
            elif parsed.path == '/api/upload-additional':
                self._handle_upload_additional()
            elif parsed.path == '/api/generate':
                self._handle_generate()
            else:
                self._json_response({'status': 'error', 'error': 'Not found'}, 404)
        except Exception as e:
            traceback.print_exc()
            self._json_response({'status': 'error', 'error': str(e)}, 500)

    def do_OPTIONS(self):
        self.send_response(200)
        self._set_cors()
        self.send_header('Content-Length', '0')
        self.end_headers()

    # --- Upload (다중 파일) ---

    def _handle_upload(self):
        """다중 파일 업로드 → 파싱 → 세금 계산"""
        content_type = self.headers.get('Content-Type', '')
        if 'multipart/form-data' not in content_type:
            self._json_response({'status': 'error', 'error': 'multipart/form-data required'}, 400)
            return

        boundary = ''
        for part in content_type.split(';'):
            part = part.strip()
            if part.startswith('boundary='):
                boundary = part.split('=', 1)[1].strip('"')
                break

        body = self.rfile.read(int(self.headers.get('Content-Length', 0)))
        files = self._parse_multipart_files(body, boundary)

        if not files:
            self._json_response({'status': 'error', 'error': '파일이 업로드되지 않았습니다'}, 400)
            return

        os.makedirs(self.UPLOAD_DIR, exist_ok=True)
        all_trades = []
        parsed_files = []
        parser = AutoDetectParser()

        for fname, fdata in files:
            ext = Path(fname).suffix.lower() if fname else '.pdf'
            tmp_path = os.path.join(self.UPLOAD_DIR, f'{uuid.uuid4().hex}{ext}')
            try:
                with open(tmp_path, 'wb') as f:
                    f.write(fdata)
                result = parser.parse(tmp_path)
                all_trades.extend(result['trades'])
                parsed_files.append({
                    'filename': fname,
                    'broker': result['broker'],
                    'trade_count': result['count'],
                })
            except Exception as e:
                parsed_files.append({
                    'filename': fname,
                    'broker': 'Error',
                    'trade_count': 0,
                    'error': str(e),
                })
            finally:
                try: os.unlink(tmp_path)
                except: pass

        cleanup_sessions()
        sid = uuid.uuid4().hex[:12]
        tax = calculate_tax(all_trades)
        sessions[sid] = {
            'trades': all_trades,
            'files': parsed_files,
            'created': datetime.now().isoformat(),
        }

        self._json_response({
            'status': 'success',
            'session_id': sid,
            'trade_count': len(all_trades),
            'trades': all_trades,
            'tax': tax,
            'files': parsed_files,
        })

    def _handle_upload_additional(self):
        """기존 세션에 추가 파일 업로드"""
        content_type = self.headers.get('Content-Type', '')
        if 'multipart/form-data' not in content_type:
            self._json_response({'status': 'error', 'error': 'multipart/form-data required'}, 400)
            return

        boundary = ''
        for part in content_type.split(';'):
            part = part.strip()
            if part.startswith('boundary='):
                boundary = part.split('=', 1)[1].strip('"')
                break

        body = self.rfile.read(int(self.headers.get('Content-Length', 0)))
        files = self._parse_multipart_files(body, boundary)
        sid = self._extract_field(body, boundary, 'session_id')

        if not sid or sid not in sessions:
            self._json_response({'status': 'error', 'error': '세션을 찾을 수 없습니다'}, 400)
            return

        os.makedirs(self.UPLOAD_DIR, exist_ok=True)
        parser = AutoDetectParser()
        new_trades = []

        for fname, fdata in files:
            ext = Path(fname).suffix.lower() if fname else '.pdf'
            tmp_path = os.path.join(self.UPLOAD_DIR, f'{uuid.uuid4().hex}{ext}')
            try:
                with open(tmp_path, 'wb') as f:
                    f.write(fdata)
                result = parser.parse(tmp_path)
                new_trades.extend(result['trades'])
                sessions[sid]['files'].append({
                    'filename': fname,
                    'broker': result['broker'],
                    'trade_count': result['count'],
                })
            finally:
                try: os.unlink(tmp_path)
                except: pass

        sessions[sid]['trades'].extend(new_trades)
        all_trades = sessions[sid]['trades']
        tax = calculate_tax(all_trades)

        self._json_response({
            'status': 'success',
            'session_id': sid,
            'trade_count': len(all_trades),
            'trades': all_trades,
            'tax': tax,
            'files': sessions[sid]['files'],
            'new_trades_count': len(new_trades),
        })

    # --- Generate ---

    def _handle_generate(self):
        """홈택스 엑셀 생성"""
        data = self._read_json()
        sid = data.get('session_id')
        auto_fill = data.get('auto_fill_buy_dates', False)

        if sid and sid in sessions:
            trades = sessions[sid]['trades']
        elif 'trades' in data:
            trades = data['trades']
        else:
            self._json_response({'status': 'error', 'error': 'session_id 필요'}, 400)
            return

        excel_bytes = generate_hometax_excel(trades, auto_fill=auto_fill)

        self.send_response(200)
        self._set_cors()
        self.send_header('Content-Type',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        fname = f'hometax_upload_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        self.send_header('Content-Disposition', f'attachment; filename="{fname}"')
        self.send_header('Content-Length', str(len(excel_bytes)))
        self.end_headers()
        self.wfile.write(excel_bytes)

    # --- Multipart Helpers ---

    def _parse_multipart_files(self, body, boundary):
        """다중 파일 추출"""
        boundary_bytes = f'--{boundary}'.encode()
        parts = body.split(boundary_bytes)
        files = []

        for part in parts:
            if b'Content-Disposition' not in part:
                continue
            header_end = part.find(b'\r\n\r\n')
            if header_end < 0:
                continue
            header = part[:header_end].decode('utf-8', errors='replace')
            content = part[header_end + 4:]
            if content.endswith(b'\r\n'):
                content = content[:-2]
            if content.endswith(b'--\r\n'):
                content = content[:-4]
            elif content.endswith(b'--'):
                content = content[:-2]
            if content.endswith(b'\r\n'):
                content = content[:-2]

            fn_match = re.search(r'filename="([^"]+)"', header)
            if fn_match and len(content) > 0:
                files.append((fn_match.group(1), content))

        return files

    def _extract_field(self, body, boundary, field_name):
        """multipart에서 텍스트 필드 추출"""
        boundary_bytes = f'--{boundary}'.encode()
        parts = body.split(boundary_bytes)
        for part in parts:
            if b'Content-Disposition' not in part:
                continue
            header_end = part.find(b'\r\n\r\n')
            if header_end < 0:
                continue
            header = part[:header_end].decode('utf-8', errors='replace')
            if f'name="{field_name}"' in header and 'filename=' not in header:
                content = part[header_end + 4:]
                if content.endswith(b'\r\n'):
                    content = content[:-2]
                return content.decode('utf-8', errors='replace').strip()
        return ''

    def _read_json(self):
        length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(length)
        return json.loads(body) if body else {}

    def _json_response(self, data, status=200):
        body = json.dumps(data, ensure_ascii=False, default=str).encode('utf-8')
        self.send_response(status)
        self._set_cors()
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Content-Length', str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _set_cors(self):
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')

    def _serve_file(self, filename, content_type):
        filepath = os.path.join(self.STATIC_DIR, filename)
        if os.path.exists(filepath):
            with open(filepath, 'rb') as f:
                data = f.read()
            self.send_response(200)
            self._set_cors()
            self.send_header('Content-Type', content_type)
            self.send_header('Content-Length', str(len(data)))
            self.end_headers()
            self.wfile.write(data)
        else:
            self.send_response(404)
            self.end_headers()

    def log_message(self, format, *args):
        print(f"[{datetime.now().strftime('%H:%M:%S')}] {format % args}")


# ============================================================
# Server
# ============================================================

class ReusableHTTPServer(HTTPServer):
    allow_reuse_address = True

def run_server(port=8080):
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    server = ReusableHTTPServer(('0.0.0.0', port), TaxEasyHandler)
    print(f"\n  TaxEasy Global v3.0 | http://localhost:{port}\n")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n서버 종료")
        server.server_close()

if __name__ == '__main__':
    port = int(sys.argv[1]) if len(sys.argv) > 1 else 8080
    run_server(port)
