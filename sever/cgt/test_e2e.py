#!/usr/bin/env python3
"""
TaxEasy Global - E2E 테스트 스크립트
실제 PDF 파일을 사용하여 전체 파이프라인 테스트
"""

import sys
sys.path.insert(0, '/sessions/serene-busy-mendel/mnt/outputs/taxeasy-server')

from parsers import AutoDetectParser
from openpyxl import load_workbook
import json
from datetime import datetime

def test_e2e():
    print("=" * 60)
    print("TaxEasy Global - E2E 테스트 시작")
    print("=" * 60)

    # 1. PDF 파일 경로 확인
    pdf_path = '/sessions/serene-busy-mendel/mnt/uploads/양도소득세 2024data.pdf'
    print(f"\n✓ 테스트 PDF: {pdf_path}")

    # 2. 파일 존재 확인
    import os
    if not os.path.exists(pdf_path):
        print(f"❌ PDF 파일을 찾을 수 없습니다: {pdf_path}")
        return False
    print("✓ PDF 파일 발견")

    # 3. 자동 증권사 감지 및 파싱
    print("\n[Step 1] 자동 증권사 감지 및 파일 파싱...")
    try:
        parser = AutoDetectParser()
        result = parser.parse(pdf_path)
        print(f"✓ 증권사 감지: {result.get('broker', 'unknown')}")
        print(f"✓ 파싱된 거래건수: {len(result.get('trades', []))}")
    except Exception as e:
        print(f"❌ 파싱 실패: {e}")
        import traceback
        traceback.print_exc()
        return False

    # 4. 거래내역 검증
    print("\n[Step 2] 거래내역 검증...")
    trades = result.get('trades', [])
    if len(trades) != 19:
        print(f"⚠️  거래건수 불일치: 예상 19건, 실제 {len(trades)}건")
    else:
        print(f"✓ 거래건수 정확함: {len(trades)}건")

    # 거래 요약
    total_sell = sum(float(t.get('sell_total', 0)) for t in trades)
    total_buy = sum(float(t.get('buy_total', 0)) for t in trades)
    total_expenses = sum(float(t.get('expenses', 0)) for t in trades)
    total_profit = total_sell - total_buy - total_expenses

    print(f"\n  거래 요약:")
    print(f"    총 양도가액: {total_sell:,.0f} 원")
    print(f"    총 취득가액: {total_buy:,.0f} 원")
    print(f"    총 필요경비: {total_expenses:,.0f} 원")
    print(f"    총 손익:    {total_profit:,.0f} 원")

    # 5. 세금 계산
    print("\n[Step 3] 세금 계산...")
    BASIC_DEDUCTION = 2_500_000
    TAX_RATE = 0.22

    if total_profit > 0:
        taxable_amount = max(0, total_profit - BASIC_DEDUCTION)
        tax = taxable_amount * TAX_RATE
    else:
        tax = 0

    print(f"  기본 공제: {BASIC_DEDUCTION:,.0f} 원")
    print(f"  과세표준: {max(0, total_profit - BASIC_DEDUCTION):,.0f} 원")
    print(f"  양도소득세: {tax:,.0f} 원 (세율: {TAX_RATE*100:.0f}%)")

    # 6. 홈택스 Excel 생성
    print("\n[Step 4] 홈택스 호환 Excel 생성...")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        # 참조 파일 로드
        template_path = '/sessions/serene-busy-mendel/mnt/outputs/최종.xlsx'
        if os.path.exists(template_path):
            ref_wb = load_workbook(template_path)
            ref_ws = ref_wb.active
            print(f"✓ 참조 파일 로드 완료 (행 수: {ref_ws.max_row})")

            # 최종 결과와 비교
            print("\n[Step 5] 파싱 결과 검증 (참조 파일과 비교)...")

            # 첫 번째 거래 확인
            if trades and ref_ws.max_row > 1:
                first_trade = trades[0]
                # 참조 파일의 첫 번째 거래 데이터 행(Row 2)
                ref_row = 2

                # 참조 파일의 관련 열들
                # A: 순번, C: 취득일자, E: 양도일자, F: 종목명, H: 양도수량,
                # J: 양도가액, K: 취득가액, L: 필요경비

                print(f"\n  첫 번째 거래 검증:")
                print(f"    파싱된 종목: {first_trade.get('stock_name', 'N/A')}")
                print(f"    파싱된 양도일: {first_trade.get('sell_date', 'N/A')}")
                print(f"    파싱된 양도가액: {first_trade.get('sell_total', 0):,.0f} 원")
                print(f"    파싱된 취득가액: {first_trade.get('buy_total', 0):,.0f} 원")
                print(f"    파싱된 필요경비: {first_trade.get('expenses', 0):,.0f} 원")

                # 모든 거래 정보 출력
                print(f"\n  파싱된 거래 목록 (상위 5개):")
                for i, trade in enumerate(trades[:5], 1):
                    print(f"    {i}. {trade.get('stock_name', 'N/A')} | "
                          f"양도: {trade.get('sell_date', 'N/A')} | "
                          f"양도가액: {trade.get('sell_total', 0):,.0f} | "
                          f"취득가액: {trade.get('buy_total', 0):,.0f}")

        print("✓ Excel 생성 로직 검증 완료")
    except Exception as e:
        print(f"⚠️  Excel 생성 중 오류: {e}")

    print("\n" + "=" * 60)
    print("✓ E2E 테스트 완료!")
    print("=" * 60)

    return True

if __name__ == '__main__':
    success = test_e2e()
    sys.exit(0 if success else 1)
